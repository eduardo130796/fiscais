import streamlit as st
import os
from io import BytesIO
import pandas as pd
import requests
from unidecode import unidecode
import time
import io
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
# Suas credenciais OAuth 2.0
CLIENT_ID = "652191149879-n4l39h8quk3rfg4lmb2ijeb4pbm347af.apps.googleusercontent.com"
CLIENT_SECRET = "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"
REFRESH_TOKEN = "1//04eIbByO3WEVXCgYIARAAGAQSNwF-L9Irele9iUZ31ls60LWTYjDfQ2_Ac5IQqJAxme2bT7JTuoTqZIoso6dI1RgUd_In3qFT7Xo"  # Use o refresh_token que você obteve na primeira autenticação

# ID da pasta específica
PASTA_ID = "1acmSiq3Blgd301aakWQU-URAy2hxoPNS"  # Coloque o ID da sua pasta aqui

# Função para renovar o token usando o refresh_token
def renovar_token(refresh_token):
    url = "https://oauth2.googleapis.com/token"
    data = {
        'client_id': CLIENT_ID,
        'client_secret': CLIENT_SECRET,
        'refresh_token': refresh_token,
        'grant_type': 'refresh_token'
    }
    
    response = requests.post(url, data=data)
    if response.status_code == 200:
        token_info = response.json()
        access_token = token_info.get('access_token')
        return access_token
    else:
        return None
# Função para obter o token de acesso
def get_access_token():
    global ACCESS_TOKEN
    if not hasattr(get_access_token, "expires_at") or time.time() > get_access_token.expires_at:
        # Se o token tiver expirado ou não estiver definido, renove-o
        ACCESS_TOKEN = renovar_token(REFRESH_TOKEN)
        get_access_token.expires_at = time.time() + 3600  # Defina o tempo de expiração como 1 hora (3600 segundos)
    return ACCESS_TOKEN

# Função para gerar o arquivo Excel em memória
def gerar_excel(df):
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    return output


cargo_mapeamento = {
    'gestor': 'GESTOR TITULAR',
    'gestorsubstituto': 'GESTOR SUBSTITUTO',
    'fiscaladministrativo': 'FISCAL ADMINISTRATIVO',
    'fiscaladministrativosubstituto': 'FISCAL ADMINISTRATIVO SUBSTITUTO',
    'fiscaltecnico': 'FISCAL TÉCNICO',
    'fiscaltecnicosubstituto': 'FISCAL TÉCNICO SUBSTITUTO'
}

# Função para normalizar o texto (remover acentos, transformar para minúsculas, remover espaços extras e "(a)")
def normalizar_texto(texto):
    if texto:
        texto = unidecode(texto)  # Remove acentos
        texto = texto.lower()  # Transforma para minúsculas
        texto = re.sub(r'\s?\(a\)\s?', '', texto)  # Remove "(a)" e "(A)"
        texto = texto.strip()  # Remove espaços extras no início e fim
        texto = ' '.join(texto.split())  # Remove espaços extras no meio
        texto = texto.replace(" ", "")  # Remove espaços entre as palavras
    return texto

# Função para normalizar a string para facilitar a comparação
def normalizar_string(s):
    return unidecode(s).lower() if isinstance(s, str) else s.lower()

def normalizar_planilha(file):
    # Carregar o arquivo Excel corretamente
    if isinstance(file, BytesIO):  # Se for um objeto BytesIO, converta
        df = pd.read_excel(file)  # Lê o conteúdo do BytesIO como um DataFrame
    else:
        df = file  # Caso já seja um DataFrame
    
    # Agora, pode aplicar a função normalizar_string
    df_normalizado = df.apply(lambda x: normalizar_string(x) if isinstance(x, str) else x)
    return df_normalizado

# Função para baixar o arquivo do Google Drive
def baixar_arquivo_drive(file_id):
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    headers = {"Authorization": f"Bearer {get_access_token()}"}
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        return io.BytesIO(response.content)  # Retorna o arquivo como BytesIO
    else:
        raise Exception(f"Erro ao baixar o arquivo: {response.status_code}")
# Função para carregar a planilha base de uma região do Google Drive
# Função para atualizar a planilha no Google Drive
def atualizar_planilha_drive(file_id, arquivo_processado):
    url = f"https://www.googleapis.com/upload/drive/v3/files/{file_id}?uploadType=media"
    headers = {
        "Authorization": f"Bearer {get_access_token()}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"  # Definindo o tipo do arquivo
    }
    
    # O conteúdo do arquivo processado será enviado
    response = requests.patch(url, headers=headers, data=arquivo_processado)

    if response.status_code == 200:
        print(f"✅ Arquivo {file_id} atualizado com sucesso!")
        return True
    else:
        print(f"❌ Erro ao atualizar o arquivo: {response.status_code}")
        print(response.text)  # Adiciona mais detalhes sobre o erro
        return False

def formatar_contrato(numero_contrato):
    """Garante que o número do contrato tenha o formato 0000/0000"""
    partes = numero_contrato.split("/")
    if len(partes) == 2:
        return f"{partes[0].zfill(4)}/{partes[1].zfill(4)}"
    return numero_contrato  # Retorna como está se não for nesse formato

# Função para processar a tabela HTML e criar o dicionário de nomes e cargos
def processar_tabela_html(html):
    soup = BeautifulSoup(html, 'html.parser')
    rows = soup.find_all('tr')
    dicionario_nomes_cargos = {
        'GESTOR TITULAR': '',
        'GESTOR SUBSTITUTO': '',
        'FISCAL ADMINISTRATIVO': '',
        'FISCAL ADMINISTRATIVO SUBSTITUTO': '',
        'FISCAL TÉCNICO': '',
        'FISCAL TÉCNICO SUBSTITUTO': ''
    }
    portaria = soup.find(attrs={"class": "Texto_Centralizado_Maiusculas"}).get_text(strip=True)
    portaria = portaria.split(",")[0]
    contrato = soup.find(attrs={"class": "Texto_Ementa"}).get_text(strip=True)
    contrato = re.findall(r"nº\s([\d/]+)", contrato)[1]
    contrato = formatar_contrato(contrato)
    for row in rows[1:]:  # Começa da segunda linha, que contém os dados
        cells = row.find_all('td')
        try:
            if len(cells) >= 3:
                unidade = cells[0].get_text(strip=True)
                unidade = re.sub(r"^DPU", "", unidade).strip()  
                nome = cells[1].get_text(strip=True)
                cargo_extraido = cells[2].get_text(strip=True)
                cargo_normalizado = normalizar_texto(cargo_extraido)
                cargo = None
                for chave, valor in cargo_mapeamento.items():
                    if cargo_normalizado == chave:
                        cargo = valor
                        break
                if cargo:
                    dicionario_nomes_cargos[cargo] = nome
        except Exception as e:
            print(f"Erro ao processar a linha: {e}")
    
    hrs = soup.find_all('hr')
    if len(hrs) > 1:
        segundo_hr = hrs[2]
    # Encontrar a tabela logo após o <hr>
        table = segundo_hr.find_next('table')

        tds = table.find_all('td')
        if len(tds) >= 2:
            processo = tds[0].get_text(strip=True)
            documento = tds[1].get_text(strip=True)
            documento = re.match(r"^\d+", documento).group() 

    return dicionario_nomes_cargos, portaria, contrato, unidade, processo, documento

# Função para buscar os dados de uma pessoa pelo nome
def buscar_dados(nome):
    nome = nome.strip()  # Remover espaços e normalizar
    nome_normalizado = normalizar_string(nome)  # Normalizar a string
    df_normalizado = normalizar_planilha(df)  # Normalizar a planilha
    for coluna in df_normalizado.columns:
        for idx, value in df_normalizado[coluna].items():
            if normalizar_string(str(value)) == nome_normalizado:
                dados_coluna = df.columns.get_loc(coluna) + 1
                return df_normalizado.iloc[idx, dados_coluna]
    return None

def atualizar_planilha(df, contrato, unidade, nomes_e_cargos, portaria, processo, documento):
    df_original = df.copy()
    erro_detectado = False  # Copiar a planilha original para comparações futuras
    df["Nº CONTRATO"] = df["Nº CONTRATO"].apply(formatar_contrato) 
    linha_existente = df[(df['Nº CONTRATO'] == contrato) & (df['UNIDADE'] == unidade)]
    
    if not linha_existente.empty:
        # Se o contrato e a unidade existem, realiza as atualizações
        for cargo, novo_nome in nomes_e_cargos.items():
            dados_novos = buscar_dados(novo_nome)
            if dados_novos is not None:
                dados_coluna = df.columns.get_loc(cargo) + 1
                df.at[linha_existente.index[0], df.columns[dados_coluna]] = dados_novos
            else:
                dados_coluna = df.columns.get_loc(cargo) + 1
                df.at[linha_existente.index[0], df.columns[dados_coluna]] = ""
                st.warning(f"🚨 **Pessoa não localizada**: A pessoa '{novo_nome}' não foi encontrada. Você precisa atuar!")
        
        # Comparação de nome atual e novo
        for cargo, novo_nome in nomes_e_cargos.items():
            nome_atual = linha_existente.iloc[0, df.columns.get_loc(cargo)].strip()
            if nome_atual != novo_nome:
                df.at[linha_existente.index[0], cargo] = novo_nome
        coluna_portaria = "Nº PORTARIA (Nº SEI)"
        df.at[linha_existente.index[0], coluna_portaria] = f"{portaria} ({documento})"

    else:
        # Caso o contrato exista mas a unidade não corresponda
        linha_contrato_existente = df[df['Nº CONTRATO'] == contrato]
        if not linha_contrato_existente.empty:
            st.warning("🚨 **Contrato já cadastrado**, mas a **unidade não confere**. Verifique a grafia da unidade.")
            erro_detectado = True
        else:
            # Caso o contrato e a unidade não existam, criar uma nova linha
            nova_linha = {'UNIDADE': unidade, 'Nº CONTRATO': contrato, 'Nº PROCESSO':processo,}
            for cargo, nome in nomes_e_cargos.items():
                nova_linha[cargo] = nome
                dados_novos = buscar_dados(nome)
                if dados_novos is not None:
                    dados_coluna = df.columns.get_loc(cargo) + 1
                    nova_linha[df.columns[dados_coluna]] = dados_novos
                else:
                    dados_coluna = df.columns.get_loc(cargo) + 1
                    nova_linha[df.columns[dados_coluna]] = ""
                    st.warning(f"🚨 **Pessoa não localizada**: A pessoa '{nome}' não foi encontrada. Você precisa atuar!")
            
            nova_linha["Nº PORTARIA (Nº SEI)"] = f"{portaria} ({documento})"
            
            nova_linha_df = pd.DataFrame([nova_linha])
            df = pd.concat([df, nova_linha_df], ignore_index=True)
    
    return df, df_original, erro_detectado

def formatar_planilha(arquivo):
    """Aplica formatação à planilha antes de enviar para o Google Drive."""
    wb = load_workbook(arquivo)
    ws = wb.active

    # Definição de estilos
    negrito = Font(bold=True)
    alinhamento_central = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Quebra de linha ativada
    borda_fina = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    fundo_cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Cor cinza no cabeçalho

    # Definir larguras personalizadas
    largura_A_D = 25  # Largura fixa para colunas A até D
    largura_E_Q = 40 # Largura fixa para colunas E até Q
    altura_padrao = 75  # Altura fixa para todas as linhas

    # Aplicar formatação no cabeçalho
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = negrito
        cell.alignment = alinhamento_central
        cell.border = borda_fina
        cell.fill = fundo_cinza  # Aplicar fundo cinza no cabeçalho

    # Aplicar bordas, alinhamento e quebra de linha no restante da planilha
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alinhamento_central  # Alinhamento central e quebra de linha
            cell.border = borda_fina

    # Definir larguras personalizadas para as colunas
    for col_num in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        if col_num <= 4:  # Colunas A até D
            ws.column_dimensions[col_letter].width = largura_A_D
        elif 5 <= col_num <= 17:  # Colunas E até Q
            ws.column_dimensions[col_letter].width = largura_E_Q

    # Aplicar altura fixa para todas as linhas
    for row_num in range(2, ws.max_row + 1):
        ws.row_dimensions[row_num].height = altura_padrao
    # Salvar as alterações
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Função para mostrar as diferenças entre antes e depois (somente as linhas alteradas)
def mostrar_diferencas(df_original, df_atualizado):
    linhas_alteradas = []
    novas_linhas = []
    for i in range(len(df_atualizado)):
        if i >= len(df_original):
            novas_linhas.append(i)
        else:
            linha_original = df_original.iloc[i]
            linha_atualizada = df_atualizado.iloc[i]
            if not linha_original.equals(linha_atualizada):
                linhas_alteradas.append(i)

    detalhes_das_alteracoes = []
    for i in linhas_alteradas:
        row_diff = {}
        for col in df_atualizado.columns:
            if df_original[col].iloc[i] != df_atualizado[col].iloc[i]:
                detalhe = f"""
                <div style="margin-bottom: 10px; border: 1px solid #ddd; border-radius: 5px; padding: 10px; background-color: #f9f9f9;">
                    <strong>{col}:</strong> 
                    <span style="color: red; text-decoration: line-through;">{df_original[col].iloc[i]}</span> 
                    <span style="color: green; font-weight: bold;">{df_atualizado[col].iloc[i]}</span>
                </div>
                """
                detalhes_das_alteracoes.append(detalhe)

    novas_linhas_detalhes = []
    for i in novas_linhas:
        detalhe = f"""
        <div style="margin-bottom: 10px; border: 1px solid #ddd; border-radius: 5px; padding: 10px; background-color: #e0f7fa;">
            <strong>Nova Linha Incluída ({i+1}):</strong> 
            {', '.join([f"<span style='color: green; font-weight: bold;'>{val}</span>" for val in df_atualizado.iloc[i].values])}
        </div>
        """
        novas_linhas_detalhes.append(detalhe)

    if detalhes_das_alteracoes or novas_linhas_detalhes:
        return detalhes_das_alteracoes, novas_linhas_detalhes, df_atualizado.iloc[linhas_alteradas], df_original.iloc[linhas_alteradas], df_atualizado.iloc[novas_linhas]
    
    return None, None, None, None, None
# Função para gerar o arquivo Excel para download

def mostrar_novas_linhas_organizadas_com_borda(novas_linhas_detalhes, novas_linhas_df):
    if novas_linhas_detalhes:
        # Organizando em duas colunas e adicionando borda
        st.markdown("""
        <style>
            .card-container {
                display: flex;
                flex-wrap: wrap;
                gap: 20px;
                justify-content: space-between;
                margin-top: 20px;
                padding: 10px;
                background-color: #f9f9f9;
                border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
            }
            .card {
                background-color: #ffffff;
                border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0, 0, 0, 0.05);
                margin: 10px;
                padding: 15px;
                width: 48%;
                border: 2px solid #ccc;
                display: flex;
                flex-direction: column;
            }
            .card h4 {
                color: #333;
                font-size: 18px;
                font-weight: bold;
                margin-bottom: 10px;
            }
            .card .field-container {
                display: grid;
                grid-template-columns: repeat(2, 1fr);
                gap: 20px;
            }
            .field-container .field {
                margin: 5px 0;
            }
            .field span {
                font-weight: bold;
                color: #4CAF50;
            }
            .faded {
                color: #777;
            }
            @media (max-width: 768px) {
                .card {
                    width: 100%;
                }
                .field-container .field {
                    width: 100%;
                }
            }
        </style>
        """, unsafe_allow_html=True)

        # Exibindo as novas linhas com borda e organizadas em duas colunas
        for idx, row in novas_linhas_df.iterrows():
            #st.markdown("<div class='card'>", unsafe_allow_html=True)
            st.markdown(f"<h4>Nova Linha Incluída - {idx + 1}</h4>", unsafe_allow_html=True)

            # Organizando as informações dentro de duas colunas com grid
            st.markdown("<div class='field-container'>", unsafe_allow_html=True)
            for i, (col, value) in enumerate(row.items()):
                col_title = col.replace('_', ' ').title()  # Formatando o título da coluna
                if pd.isna(value):
                    st.markdown(f"<div class='field'><span>{col_title}:</span> <span class='faded'>Nenhuma informação</span></div>", unsafe_allow_html=True)
                else:
                    st.markdown(f"<div class='field'><span>{col_title}:</span> {value}</div>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)



# Streamlit interface
st.set_page_config(page_title="📝 Atualizador de Planilha por Região", page_icon="📑", layout="wide")
st.title('📑 Atualizador de Planilha por Região')
st.markdown("""  
    **Bem-vindo ao atualizador de planilhas por região!**  
    Este aplicativo permite atualizar os dados de fiscais em uma planilha com base nas informações extraídas de uma tabela HTML para diferentes regiões.
    
    🛠️ **Passos:**  
    1. Selecione a região e cole o HTML da tabela.  
    2. Atualize a planilha para cada região e baixe os arquivos atualizados.
""", unsafe_allow_html=True)
# Organizando as regiões horizontalmente usando colunas
col1, col2, col3, col4, col5 = st.columns(5)

# Checkboxes para selecionar as regiões, organizadas horizontalmente
selecionadas = [
    col1.checkbox('Centro-Oeste'),
    col2.checkbox('Nordeste'),
    col3.checkbox('Sudeste'),
    col4.checkbox('Sul'),
    col5.checkbox('Norte')
]
# Se alguma região for selecionada, exibe o campo para o HTML correspondente
for i, regiao in enumerate(['Centro-Oeste', 'Nordeste', 'Sudeste', 'Sul', 'Norte']):
    if selecionadas[i]:
        with st.expander(f"🔧 Configurações para {regiao}"):
            html_input = st.text_area(f"HTML da Região: {regiao}", height=100)
        
        # Se o HTML for fornecido, processar e atualizar a planilha
        if html_input:
            # Carregar planilha do Google Drive com base na região
            file_id = {
                "Centro-Oeste": "ID_PLANILHA_CO", 
                "Nordeste": "1ZF30tvr_Tb5AcEMv2nR0drGTFWz95Ysn", 
                "Sudeste": "ID_PLANILHA_SE", 
                "Sul": "ID_PLANILHA_SU", 
                "Norte": "ID_PLANILHA_NO"
            }[regiao]

            df = baixar_arquivo_drive(file_id)
            df = pd.read_excel(df)

            if df is not None:
                dicionario, portaria, contrato, unidade, processo, documento = processar_tabela_html(html_input)
                st.subheader("📄 Informações:")
                st.write(f"**Contrato**: {contrato}")
                st.write(f"**Unidade**: {unidade}")
                st.write(dicionario)


                if st.button(f"📤 Atualizar Planilha - {regiao}"):
                    df_atualizado, df_original, erro_detectado = atualizar_planilha(df, contrato, unidade, dicionario, portaria, processo, documento)

                    if erro_detectado:
                        st.warning(f"🚨 **Erro detectado**: O contrato já está registrado, mas a unidade não confere. Por favor, verifique.")
                    else:
                        arquivo_processado = gerar_excel(df_atualizado)
                        #arquivo_processado = converter_df_para_excel(df_atualizado)
                        arquivo_formatado = formatar_planilha(arquivo_processado)

                        # Enviar para o Google Drive com o arquivo formatado
                        atualizar_planilha_drive(file_id, arquivo_formatado.getvalue())

                        #atualizar_planilha_drive(file_id, arquivo_processado)
                        st.success(f"✔️ Planilha da região {regiao} atualizada com sucesso!")

                        

                        detalhes, novas_linhas_detalhes, df_alteradas_atual, df_alteradas_original, novas_linhas_df = mostrar_diferencas(df_original, df_atualizado)

                        # Exibir novas linhas incluídas com estilo
                        if novas_linhas_detalhes:
                            mostrar_novas_linhas_organizadas_com_borda(novas_linhas_detalhes, novas_linhas_df)
                            st.dataframe(novas_linhas_df, use_container_width=True)
                        else:
                            # Mostrar as diferenças entre a planilha original e a atualizada
                            if detalhes:
                                st.subheader("🔄 Diferenças - Linhas Alteradas:")
                                for detalhe in detalhes:
                                    st.markdown(detalhe, unsafe_allow_html=True)

                                col1, col2 = st.columns(2)
                                with col1:
                                    st.subheader("📝 Planilha Original - Linhas Alteradas")
                                    st.dataframe(df_alteradas_original, use_container_width=True)

                                with col2:
                                    st.subheader("🔄 Planilha Atualizada - Linhas Alteradas")
                                    st.dataframe(df_alteradas_atual, use_container_width=True)

                            else:
                                st.write("🔔 Não há alterações nas linhas.")

                        # Adicionar botão de download
                        excel_file = gerar_excel(df_atualizado)
                        st.download_button(
                            label="💾 Baixar Planilha Atualizada",
                            data=excel_file,
                            file_name="planilha_atualizada.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )